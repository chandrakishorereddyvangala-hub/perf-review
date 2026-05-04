import os
import re
import json
import time
import threading
from io import BytesIO
import gspread
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file

load_dotenv()  # reads .env locally; ignored on Vercel (uses dashboard env vars)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "perfreview_secret_2024")

# ── Google Sheets config ─────────────────────────────────────────────────────
SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID", "")
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

_gc = None
_gc_lock = threading.Lock()

def get_client():
    global _gc
    if _gc is not None:          # fast path — no lock needed once initialised
        return _gc
    with _gc_lock:               # slow path — only one thread initialises
        if _gc is not None:
            return _gc
        raw = os.environ.get("GOOGLE_CREDENTIALS", "")
        if raw:
            info = json.loads(raw)
            creds = Credentials.from_service_account_info(info, scopes=SCOPES)
            _gc = gspread.authorize(creds)
        elif os.path.exists("credentials.json"):
            _gc = gspread.service_account(filename="credentials.json")
        else:
            raise RuntimeError("No Google credentials found.")
    return _gc

def get_spreadsheet():
    return get_client().open_by_key(SPREADSHEET_ID)


# ── Simple 30-second in-memory cache ─────────────────────────────────────────
# Keeps the app snappy without hammering the Sheets API on every page load.
# Any change you make in the spreadsheet reflects in the app within 30 seconds.
_cache: dict = {}
_cache_lock = threading.Lock()
CACHE_TTL = 300  # seconds

def _cached(key, loader_fn):
    now = time.time()
    # Fast path — no lock, cache is warm
    entry = _cache.get(key)
    if entry and (now - entry["ts"]) < CACHE_TTL:
        return entry["data"]
    # Slow path — only one thread calls the loader for this key
    with _cache_lock:
        entry = _cache.get(key)          # recheck under lock
        if entry and (now - entry["ts"]) < CACHE_TTL:
            return entry["data"]
        data = loader_fn()
        _cache[key] = {"data": data, "ts": now}
        return data

def bust_cache(*keys):
    for k in keys:
        _cache.pop(k, None)


# ── Users — manage logins here directly ──────────────────────────────────────
# To add a user:    "username": {"password": "yourpass", "role": "lead"}
# To add director:  "username": {"password": "yourpass", "role": "director"}
# To remove:        delete the line
# Usernames are case-insensitive (always stored lowercase)
USERS = {
    "Chandra":   {"password": "pass1234",    "role": "lead"},
    "Uma":       {"password": "pass123",    "role": "lead"},
    "Vinoth":    {"password": "pass123",    "role": "lead"},
    "Tejas":     {"password": "pass123",    "role": "lead"},
    "Suresh":    {"password": "pass123",    "role": "lead"},
    "Aishwarya": {"password": "pass123",    "role": "lead"},
    "Dheeraj":   {"password": "pass123",    "role": "lead"},
    "Drushya":   {"password": "pass123",    "role": "lead"},
    "NaveenR":   {"password": "pass123",    "role": "lead"},
    "Naveen":    {"password": "pass123",    "role": "director"},
}
LEADS = [u for u, d in USERS.items() if d["role"] == "lead"]
DIRECTORS = [u for u, d in USERS.items() if d["role"] == "director"]

# ── Constants ─────────────────────────────────────────────────────────────────
RATING_CATEGORIES = [
    "Collaboration & Communication",
    "Continuous Improvement & Learning",
    "Compliance & Professional Conduct",
    "Ownership & Value Addition",
    "Stakeholder Satisfaction",
    "Exceptional Contribution",
    "Feedback & Improvement",
    "Work Discipline & Professional Practice",
    "SLA & Deadline Adherence",
    "Automation & Tools",
    "Quality Metrics & Reporting",
]
REV_HEADERS = (
    ["employee", "status"]
    + RATING_CATEGORIES
    + ["notes", "lead_comments", "comments", "shared_with"]
)


# ── Sheet initialisation (once on first request) ──────────────────────────────
_sheet_ready = False

def _migrate_lead_sheet(ws):
    """Rewrite ws to match REV_HEADERS, preserving any column that kept its name."""
    all_data = ws.get_all_values()
    if not all_data:
        ws.update('A1', [REV_HEADERS])
        return
    old_headers = all_data[0]
    if old_headers == REV_HEADERS:
        return  # already current
    rows = all_data[1:]
    new_data = [REV_HEADERS]
    for row in rows:
        old_rec = _zip(old_headers, row)
        if not old_rec.get("employee"):
            continue  # skip blank rows
        new_row = [old_rec.get(h, "") for h in REV_HEADERS]
        new_data.append(new_row)
    ws.clear()
    time.sleep(0.3)
    ws.update('A1', new_data)

def _init_sheets():
    global _sheet_ready
    if _sheet_ready:
        return
    sh = get_spreadsheet()
    existing = {ws.title for ws in sh.worksheets()}

    # users tab — the master list of logins & roles
    if "users" not in existing:
        ws = sh.add_worksheet("users", rows=100, cols=4)
        time.sleep(0.8)
        ws.append_row(["username", "password", "role"])
        defaults = [
            ("chandra",   "pass123", "lead"),
            ("uma",       "pass123", "lead"),
            ("vinoth",    "pass123", "lead"),
            ("tejas",     "pass123", "lead"),
            ("suresh",    "pass123", "lead"),
            ("aishwarya", "pass123", "lead"),
            ("dheeraj",   "pass123", "lead"),
            ("drushya",   "pass123", "lead"),
            ("naveen",    "pass123", "director"),
        ]
        for row in defaults:
            ws.append_row(list(row))
            time.sleep(0.2)

    # org tab — resource → lead mapping
    if "org" not in existing:
        ws = sh.add_worksheet("org", rows=200, cols=5)
        time.sleep(0.8)
        ws.append_row(["employee", "lead", "role"])

    # notifications tab
    if "notifications" not in existing:
        ws = sh.add_worksheet("notifications", rows=500, cols=4)
        time.sleep(0.8)
        ws.append_row(["recipient", "message", "timestamp", "is_read"])

    # migrate all lead sheets to current REV_HEADERS (reuse worksheets list)
    all_ws = {ws.title.lower(): ws for ws in sh.worksheets()}
    for lead in LEADS:
        ws = all_ws.get(lead.lower())
        if ws:
            try:
                _migrate_lead_sheet(ws)
                time.sleep(0.3)
            except Exception:
                pass

    _sheet_ready = True




# ── Dynamic loaders (everything comes from the sheet) ────────────────────────

def load_org():
    """Load org structure from the 'org' sheet. Cached 30 s."""
    def _load():
        sh = get_spreadsheet()
        try:
            ws = sh.worksheet("org")
        except gspread.WorksheetNotFound:
            return {}, {}
        headers, rows = _parse_ws(ws)
        emp_info, lead_emps = {}, {}
        for row in rows:
            rec = _zip(headers, row)
            emp  = rec.get("employee", "").strip()
            lead = rec.get("lead", "").strip().lower()
            role = rec.get("role", "Employee").strip() or "Employee"
            if emp and lead:
                emp_info[emp] = {"lead": lead, "role": role}
                lead_emps.setdefault(lead, []).append(emp)
        return emp_info, lead_emps
    return _cached("org", _load)


# ── Sheet helpers ─────────────────────────────────────────────────────────────

def _sheets_read(fn, *args, **kwargs):
    """Call a Sheets API read with up to 3 retries on 429 rate-limit errors."""
    delay = 1.5
    for attempt in range(3):
        try:
            return fn(*args, **kwargs)
        except gspread.exceptions.APIError as e:
            if "429" in str(e) and attempt < 2:
                time.sleep(delay * (attempt + 1))
                continue
            raise

def _parse_ws(ws):
    data = _sheets_read(ws.get_all_values)
    if not data:
        return [], []
    return data[0], data[1:]

def _zip(headers, row):
    padded = row + [""] * max(0, len(headers) - len(row))
    return dict(zip(headers, padded))

def _to_rec(headers, row):
    rec = _zip(headers, row)
    for field in ["comments", "shared_with"]:
        try:
            rec[field] = json.loads(rec.get(field) or "[]")
        except Exception:
            rec[field] = []
    for cat in RATING_CATEGORIES:
        try:
            rec[cat] = float(rec.get(cat) or 0)
        except (ValueError, TypeError):
            rec[cat] = 0.0
    rec["notes"] = _clean_notes(rec.get("notes", ""))
    try:
        lc = rec.get("lead_comments", "")
        rec["lead_comments"] = json.loads(lc) if lc else {}
    except Exception:
        rec["lead_comments"] = {}
    return rec

def _get_ws_ci(sh, name):
    """Case-insensitive worksheet lookup. Raises WorksheetNotFound if missing."""
    for ws in sh.worksheets():
        if ws.title.lower() == name.lower():
            return ws
    raise gspread.WorksheetNotFound(name)

def _ensure_lead_sheet(sh, lead):
    """Get or create a lead's worksheet on demand (case-insensitive match).
    Also syncs the header row if categories have changed."""
    try:
        ws = _get_ws_ci(sh, lead)
        if ws.row_values(1) != REV_HEADERS:
            ws.update('A1', [REV_HEADERS])
        return ws
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(lead, rows=200, cols=len(REV_HEADERS))
        time.sleep(0.8)
        ws.append_row(REV_HEADERS)
        return ws

def _clean_notes(raw):
    """Strip legacy __TP__....__END__ encoding — return plain text only."""
    if not raw:
        return ""
    if raw.startswith("__TP__"):
        try:
            return raw.split("__END__", 1)[1]
        except Exception:
            return ""
    return raw

def compute_avg(review):
    ratings = [review.get(c, 0) or 0 for c in RATING_CATEGORIES]
    return round(sum(ratings) / len(ratings), 1) if any(ratings) else 0


def _update_password_in_file(username, new_password):
    filepath = os.path.abspath(__file__)
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    pattern = r'("' + re.escape(username) + r'":\s*\{"password":\s*)"[^"]*"'
    escaped_pw = new_password.replace('\\', '\\\\').replace('"', '\\"')
    new_content = re.sub(pattern, r'\1"' + escaped_pw + '"', content)
    if new_content != content:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(new_content)


def load_notifications(lead):
    def _load():
        sh = get_spreadsheet()
        try:
            ws = sh.worksheet("notifications")
        except gspread.WorksheetNotFound:
            return []
        headers, rows = _parse_ws(ws)
        result = []
        for row in rows:
            rec = _zip(headers, row)
            if rec.get("recipient", "").lower() == lead.lower():
                result.append(rec)
        return result
    return _cached(f"notif_{lead}", _load)


def add_notification(recipient, message):
    try:
        sh = get_spreadsheet()
        try:
            ws = sh.worksheet("notifications")
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet("notifications", rows=500, cols=4)
            time.sleep(0.8)
            ws.append_row(["recipient", "message", "timestamp", "is_read"])
        timestamp = datetime.now().strftime("%b %d, %Y %H:%M")
        ws.append_row([recipient, message, timestamp, "false"])
        bust_cache(f"notif_{recipient}")
    except Exception:
        pass


def mark_notifications_read(lead):
    try:
        sh = get_spreadsheet()
        ws = sh.worksheet("notifications")
        all_data = ws.get_all_values()
        if not all_data:
            return
        headers = all_data[0]
        recip_idx = headers.index("recipient")
        read_idx = headers.index("is_read")
        for i, row in enumerate(all_data[1:], start=2):
            padded = row + [""] * max(0, len(headers) - len(row))
            if padded[recip_idx].lower() == lead.lower() and padded[read_idx] != "true":
                ws.update_cell(i, read_idx + 1, "true")
        bust_cache(f"notif_{lead}")
    except Exception:
        pass


def add_employee_to_org(lead, emp_name, emp_role):
    sh = get_spreadsheet()
    try:
        ws = sh.worksheet("org")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet("org", rows=200, cols=3)
        time.sleep(0.8)
        ws.append_row(["employee", "lead", "role"])
    headers, rows = _parse_ws(ws)
    for row in rows:
        rec = _zip(headers, row)
        if rec.get("employee", "").strip().lower() == emp_name.strip().lower():
            return False, "A team member with that name already exists"
    ws.append_row([emp_name, lead, emp_role])
    bust_cache("org")
    return True, "Added"


def remove_employee_from_org(lead, emp_name):
    sh = get_spreadsheet()
    try:
        ws = sh.worksheet("org")
    except gspread.WorksheetNotFound:
        return False, "Org sheet not found"
    all_data = ws.get_all_values()
    if not all_data:
        return False, "Empty sheet"
    headers = all_data[0]
    try:
        emp_idx = headers.index("employee")
        lead_idx = headers.index("lead")
    except ValueError:
        return False, "Invalid sheet structure"
    for i, row in enumerate(all_data[1:], start=2):
        padded = row + [""] * max(0, len(headers) - len(row))
        if (padded[emp_idx].strip().lower() == emp_name.strip().lower() and
                padded[lead_idx].strip().lower() == lead.lower()):
            ws.delete_rows(i)
            bust_cache("org")
            return True, "Removed"
    return False, "Employee not found in your team"


def load_review(lead, emp_name):
    sh = get_spreadsheet()
    try:
        ws = _get_ws_ci(sh, lead)
    except gspread.WorksheetNotFound:
        return None
    headers, rows = _parse_ws(ws)
    for row in rows:
        if row and row[0] == emp_name:
            return _to_rec(headers, row)
    return None


def load_all_lead_reviews(lead):
    """Read all reviews for a lead in one API call."""
    def _load():
        sh = get_spreadsheet()
        try:
            ws = _get_ws_ci(sh, lead)
        except gspread.WorksheetNotFound:
            return []
        headers, rows = _parse_ws(ws)
        return [_to_rec(headers, row) for row in rows if row and row[0]]
    return _cached(f"reviews_{lead.lower()}", _load)


def save_review(lead, emp_name, data):
    sh = get_spreadsheet()
    ws = _ensure_lead_sheet(sh, lead)
    all_data = ws.get_all_values()
    headers = all_data[0] if all_data else REV_HEADERS

    def serial(val):
        if isinstance(val, (list, dict)):
            return json.dumps(val)
        return val if val is not None else ""

    values = [serial(data.get(h, "")) for h in headers]
    end_col = chr(ord("A") + len(headers) - 1)

    for i, row in enumerate(all_data[1:], start=2):
        if row and row[0] == emp_name:
            ws.update(f"A{i}:{end_col}{i}", [values])  # gspread 6.x: range first
            bust_cache(f"reviews_{lead.lower()}")
            return

    ws.append_row(values)  # new employee row
    bust_cache(f"reviews_{lead.lower()}")


def get_shared_employees(lead):
    def _load():
        sh = get_spreadsheet()
        all_ws = {ws.title.lower(): ws for ws in sh.worksheets()}
        shared = []
        for owner_lead in LEADS:
            ol_lower = owner_lead.lower()
            if ol_lower == lead or ol_lower not in all_ws:
                continue
            headers, rows = _parse_ws(all_ws[ol_lower])
            try:
                sw_idx = headers.index("shared_with")
            except ValueError:
                continue
            for row in rows:
                padded = row + [""] * max(0, len(headers) - len(row))
                try:
                    sw = json.loads(padded[sw_idx] or "[]")
                except Exception:
                    sw = []
                if lead in [s.lower() for s in sw] and padded[0]:
                    rev = _to_rec(headers, row)
                    shared.append({
                        "emp": padded[0],
                        "owner_lead": owner_lead,
                        "review": rev,
                        "avg_rating": compute_avg(rev),
                    })
        return shared
    return _cached(f"shared_{lead}", _load)


def _parse_notif_link(message):
    """Derive a review URL from a share notification message."""
    m = re.search(r"shared (.+?)'s review", message)
    return f"/review/{m.group(1)}" if m else ""


# ── Error handlers ───────────────────────────────────────────────────────────

@app.errorhandler(gspread.exceptions.APIError)
def handle_sheets_error(e):
    is_rate_limit = "429" in str(e)
    msg = ("The server is under load right now. Please wait a few seconds and refresh."
           if is_rate_limit else
           "A data error occurred. Please try again.")
    if request.path.startswith("/api/"):
        status = 503 if is_rate_limit else 500
        return jsonify({"ok": False, "error": msg}), status
    flash(msg)
    return redirect(url_for("dashboard") if "lead" in session else url_for("login"))

@app.errorhandler(500)
def handle_500(e):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": "An unexpected error occurred. Please try again."}), 500
    flash("Something went wrong. Please try again.")
    return redirect(url_for("dashboard") if "lead" in session else url_for("login"))


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/", methods=["GET", "POST"])
def login():
    if "lead" in session:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username_input = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        # Case-insensitive username match
        matched_key = next((k for k in USERS if k.lower() == username_input.lower()), None)
        user = USERS.get(matched_key) if matched_key else None
        if user and user["password"] == password:
            session["lead"] = matched_key.lower()
            session["role"] = user["role"]
            return redirect(url_for("dashboard"))
        flash("Invalid credentials. Please try again.")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
def dashboard():
    if "lead" not in session:
        return redirect(url_for("login"))

    if session.get("role") == "director":
        _, lead_emps = load_org()
        lead_tiles = []
        for lead in LEADS:
            reviews = load_all_lead_reviews(lead)
            rev_map = {r["employee"]: r for r in reviews}
            emps = lead_emps.get(lead.lower(), [])
            rated, total_avg = 0, 0.0
            status_counts = {"Pending": 0, "In Progress": 0, "Completed": 0}
            for emp in emps:
                rev = rev_map.get(emp)
                if rev:
                    avg = compute_avg(rev)
                    if avg:
                        total_avg += avg
                        rated += 1
                    s = rev.get("status", "Pending")
                    status_counts[s] = status_counts.get(s, 0) + 1
            lead_tiles.append({
                "name": lead,
                "emp_count": len(emps),
                "team_avg": round(total_avg / rated, 1) if rated else 0,
                "status_counts": status_counts,
            })
        return render_template("director_dashboard.html", lead_tiles=lead_tiles)

    lead = session["lead"]
    emp_info, lead_emps = load_org()
    reviews = load_all_lead_reviews(lead)
    rev_map = {r["employee"]: r for r in reviews}
    employees = []
    for emp in lead_emps.get(lead, []):
        rev = rev_map.get(emp)
        employees.append({
            "name": emp,
            "info": emp_info.get(emp, {"role": "Employee", "lead": lead}),
            "status": rev.get("status", "Pending") if rev else "Pending",
            "avg_rating": compute_avg(rev) if rev else 0,
        })
    shared = get_shared_employees(lead)
    return render_template("dashboard.html", lead=lead, employees=employees, shared=shared)


def _tier(avg):
    if avg == 0:    return "unrated",  "Not Rated",    "Pending"
    if avg >= 4.5:  return "top",      "Top Performer","High Priority"
    if avg >= 3.5:  return "high",     "High",         "Priority"
    if avg >= 2.5:  return "standard", "Standard",     "Standard"
    return              "review",  "Needs Review", "On Hold"


def _build_rankings():
    """Shared data builder for rankings view and Excel export."""
    emp_info, lead_emps = load_org()
    all_employees = []
    for lead in LEADS:
        reviews = load_all_lead_reviews(lead)
        rev_map = {r["employee"]: r for r in reviews}
        for emp in lead_emps.get(lead.lower(), []):
            rev = rev_map.get(emp) or {}
            avg = compute_avg(rev) if rev else 0
            tier, tier_label, appraisal = _tier(avg)
            all_employees.append({
                "name": emp,
                "lead": lead,
                "role": emp_info.get(emp, {}).get("role", "Employee"),
                "avg_rating": avg,
                "status": rev.get("status", "Pending"),
                "tier": tier,
                "tier_label": tier_label,
                "appraisal": appraisal,
                "ratings": {c: rev.get(c, 0) or 0 for c in RATING_CATEGORIES},
            })
    rated   = sorted([e for e in all_employees if e["avg_rating"] > 0],
                     key=lambda x: x["avg_rating"], reverse=True)
    unrated = [e for e in all_employees if e["avg_rating"] == 0]
    ranked  = rated + unrated
    for i, emp in enumerate(ranked, 1):
        emp["rank"] = i if emp["avg_rating"] > 0 else None
    return ranked


@app.route("/director/rankings")
def director_rankings():
    if "lead" not in session or session.get("role") != "director":
        flash("Access denied.")
        return redirect(url_for("dashboard"))
    ranked = _build_rankings()
    tier_counts = {t: sum(1 for e in ranked if e["tier"] == t)
                   for t in ("top", "high", "standard", "review", "unrated")}
    rated_avgs  = [e["avg_rating"] for e in ranked if e["avg_rating"] > 0]
    company_avg = round(sum(rated_avgs) / len(rated_avgs), 1) if rated_avgs else 0
    return render_template(
        "director_rankings.html",
        employees=ranked,
        leads=LEADS,
        tier_counts=tier_counts,
        company_avg=company_avg,
        total=len(ranked),
        rated_count=len([e for e in ranked if e["avg_rating"] > 0]),
    )


@app.route("/director/export_rankings")
def export_rankings():
    if "lead" not in session or session.get("role") != "director":
        flash("Access denied.")
        return redirect(url_for("dashboard"))

    ranked = _build_rankings()
    today  = datetime.now().strftime("%d %b %Y")
    fname  = f"appraisal_rankings_{datetime.now().strftime('%Y%m%d')}.xlsx"

    # ── Style constants ───────────────────────────────────────────
    TIER_BG   = {"top": "FEF3C7", "high": "D1FAE5", "standard": "DBEAFE",
                 "review": "FEE2E2", "unrated": "F3F4F6"}
    TIER_FG   = {"top": "92400E", "high": "065F46", "standard": "1E40AF",
                 "review": "991B1B", "unrated": "6B7280"}
    HDR_FILL  = PatternFill("solid", fgColor="242627")
    HDR_FONT  = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    TITLE_FONT= Font(name="Calibri", bold=True, size=14, color="242627")
    SUB_FONT  = Font(name="Calibri", size=10, color="6B7280")
    STAT_FONT = Font(name="Calibri", bold=True, size=11)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="D1D5DB")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════
    # Sheet 1 — Full Rankings
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Rankings"
    ws1.sheet_view.showGridLines = False

    # Title block
    ws1.merge_cells("A1:S1")
    ws1["A1"] = "NFORCE ONE — Appraisal Rankings"
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = LEFT
    ws1.row_dimensions[1].height = 28

    ws1["A2"] = f"Exported: {today}   ·   {len(ranked)} resources   ·   {len([e for e in ranked if e['avg_rating'] > 0])} rated"
    ws1["A2"].font = SUB_FONT
    ws1.merge_cells("A2:S2")
    ws1.row_dimensions[2].height = 18

    ws1.row_dimensions[3].height = 10  # spacer

    # Summary stat row
    tier_counts = {t: sum(1 for e in ranked if e["tier"] == t)
                   for t in ("top", "high", "standard", "review", "unrated")}
    rated_avgs  = [e["avg_rating"] for e in ranked if e["avg_rating"] > 0]
    company_avg = round(sum(rated_avgs) / len(rated_avgs), 1) if rated_avgs else 0

    stats = [
        ("Total Resources", len(ranked)),
        ("Rated",           len(rated_avgs)),
        ("Company Avg",     company_avg),
        ("Top Performers",  tier_counts["top"]),
        ("High",            tier_counts["high"]),
        ("Standard",        tier_counts["standard"]),
        ("Needs Review",    tier_counts["review"]),
        ("Not Rated",       tier_counts["unrated"]),
    ]
    stat_fills = ["F4F5F9","F4F5F9","F4F5F9","FEF3C7","D1FAE5","DBEAFE","FEE2E2","F3F4F6"]
    for col_idx, ((label, val), bg) in enumerate(zip(stats, stat_fills), 1):
        lc = ws1.cell(row=4, column=col_idx, value=label)
        vc = ws1.cell(row=5, column=col_idx, value=val)
        for cell in (lc, vc):
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = BORDER
            cell.alignment = CENTER
        lc.font = Font(name="Calibri", size=9, color="6B7280", bold=False)
        vc.font = Font(name="Calibri", size=13, bold=True, color="242627")
        ws1.column_dimensions[get_column_letter(col_idx)].width = 16
    ws1.row_dimensions[4].height = 16
    ws1.row_dimensions[5].height = 24

    ws1.row_dimensions[6].height = 10  # spacer

    # Column headers (row 7)
    fixed_headers = ["Rank", "Employee", "Lead", "Role", "Avg Rating", "Tier", "Appraisal Priority", "Review Status"]
    all_headers   = fixed_headers + RATING_CATEGORIES
    col_widths    = [7, 22, 14, 16, 11, 16, 18, 14] + [20] * len(RATING_CATEGORIES)

    for col_idx, (header, width) in enumerate(zip(all_headers, col_widths), 1):
        cell = ws1.cell(row=7, column=col_idx, value=header)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[7].height = 20

    # Data rows (row 8+)
    for emp in ranked:
        row_data = [
            emp["rank"] if emp["rank"] else "—",
            emp["name"],
            emp["lead"].capitalize(),
            emp["role"],
            emp["avg_rating"] if emp["avg_rating"] > 0 else "—",
            emp["tier_label"],
            emp["appraisal"],
            emp["status"],
        ] + [emp["ratings"].get(c, 0) for c in RATING_CATEGORIES]

        bg   = TIER_BG.get(emp["tier"], "FFFFFF")
        fg   = TIER_FG.get(emp["tier"], "242627")
        fill = PatternFill("solid", fgColor=bg)
        font = Font(name="Calibri", color=fg, size=10)

        r = ws1.max_row + 1
        ws1.row_dimensions[r].height = 18
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=col_idx, value=value)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = CENTER if col_idx != 2 else LEFT
            cell.border    = BORDER

    # Freeze panes below header
    ws1.freeze_panes = "A8"

    # ════════════════════════════════════════════════════════════
    # Sheet 2 — Tier Summary
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Tier Summary")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:E1")
    ws2["A1"] = "Appraisal Tier Summary"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].alignment = LEFT
    ws2.row_dimensions[1].height = 28

    ws2["A2"] = f"As of {today}"
    ws2["A2"].font = SUB_FONT
    ws2.merge_cells("A2:E2")
    ws2.row_dimensions[2].height = 16
    ws2.row_dimensions[3].height = 10

    sum_headers = ["Tier", "Count", "% of Total", "Appraisal Action", "Rating Range"]
    sum_widths  = [20, 10, 14, 22, 16]
    for col_idx, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CENTER; cell.border = BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[4].height = 20

    total_n = len(ranked) or 1
    tier_rows = [
        ("top",      "Top Performer", "High Priority",  "≥ 4.5"),
        ("high",     "High",          "Priority",       "3.5 – 4.4"),
        ("standard", "Standard",      "Standard",       "2.5 – 3.4"),
        ("review",   "Needs Review",  "On Hold",        "< 2.5"),
        ("unrated",  "Not Rated",     "Pending",        "—"),
    ]
    for tier_key, label, action, rng in tier_rows:
        cnt  = tier_counts[tier_key]
        pct  = round(cnt / total_n * 100, 1)
        bg   = TIER_BG[tier_key]; fg = TIER_FG[tier_key]
        fill = PatternFill("solid", fgColor=bg)
        font = Font(name="Calibri", color=fg, size=10, bold=True)
        r    = ws2.max_row + 1
        ws2.row_dimensions[r].height = 20
        for col_idx, val in enumerate([label, cnt, f"{pct}%", action, rng], 1):
            cell = ws2.cell(row=r, column=col_idx, value=val)
            cell.fill = fill; cell.font = font
            cell.alignment = CENTER; cell.border = BORDER

    # ════════════════════════════════════════════════════════════
    # Sheet 3 — Per-Lead Summary
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Per-Lead Summary")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:G1")
    ws3["A1"] = "Per-Lead Summary"
    ws3["A1"].font = TITLE_FONT
    ws3["A1"].alignment = LEFT
    ws3.row_dimensions[1].height = 28
    ws3.merge_cells("A2:G2")
    ws3["A2"] = f"As of {today}"
    ws3["A2"].font = SUB_FONT
    ws3.row_dimensions[2].height = 16
    ws3.row_dimensions[3].height = 10

    lead_headers = ["Lead", "Total", "Rated", "Team Avg", "Top Performers", "Needs Review", "Not Rated"]
    lead_widths  = [18, 10, 10, 12, 16, 14, 12]
    for col_idx, (h, w) in enumerate(zip(lead_headers, lead_widths), 1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CENTER; cell.border = BORDER
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.row_dimensions[4].height = 20

    for lead in LEADS:
        lead_emps = [e for e in ranked if e["lead"].lower() == lead.lower()]
        l_rated   = [e for e in lead_emps if e["avg_rating"] > 0]
        l_avg     = round(sum(e["avg_rating"] for e in l_rated) / len(l_rated), 1) if l_rated else "—"
        l_top     = sum(1 for e in lead_emps if e["tier"] == "top")
        l_review  = sum(1 for e in lead_emps if e["tier"] == "review")
        l_unrated = sum(1 for e in lead_emps if e["tier"] == "unrated")
        r = ws3.max_row + 1
        ws3.row_dimensions[r].height = 18
        for col_idx, val in enumerate([lead.capitalize(), len(lead_emps), len(l_rated),
                                       l_avg, l_top, l_review, l_unrated], 1):
            cell = ws3.cell(row=r, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor="F9FAFB" if r % 2 == 0 else "FFFFFF")
            cell.font = Font(name="Calibri", size=10, color="242627")
            cell.alignment = CENTER; cell.border = BORDER

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/director/team/<lead_name>")
def director_team(lead_name):
    if "lead" not in session or session.get("role") != "director":
        flash("Access denied.")
        return redirect(url_for("dashboard"))
    emp_info, lead_emps = load_org()
    reviews = load_all_lead_reviews(lead_name)
    rev_map = {r["employee"]: r for r in reviews}
    employees = []
    for emp in lead_emps.get(lead_name.lower(), []):
        rev = rev_map.get(emp, {})
        employees.append({
            "name": emp,
            "info": emp_info.get(emp, {"role": "Employee"}),
            "status": rev.get("status", "Pending"),
            "avg_rating": compute_avg(rev) if rev else 0,
            "ratings": {c: rev.get(c, 0) or 0 for c in RATING_CATEGORIES},
            "notes": rev.get("notes", ""),
            "lead_comments": rev.get("lead_comments", ""),
        })
    return render_template(
        "director_team.html",
        lead_name=lead_name,
        employees=employees,
        categories=RATING_CATEGORIES,
    )


@app.route("/review/<emp_name>", methods=["GET", "POST"])
def review(emp_name):
    if "lead" not in session:
        return redirect(url_for("login"))
    lead = session["lead"]
    role = session.get("role", "lead")

    if role == "director":
        flash("View-only access — use the director panel.")
        return redirect(url_for("dashboard"))

    emp_info, _ = load_org()
    info = emp_info.get(emp_name, {})
    owner_lead = info.get("lead")
    is_owner = (lead == owner_lead)

    if not is_owner:
        rev_check = load_review(owner_lead, emp_name) if owner_lead else None
        shared_with_lc = [s.lower() for s in (rev_check or {}).get("shared_with", [])]
        if not rev_check or lead not in shared_with_lc:
            flash("Access denied.")
            return redirect(url_for("dashboard"))

    review_data = load_review(owner_lead, emp_name) if owner_lead else {}
    if not review_data:
        review_data = {
            "employee": emp_name, "status": "Pending",
            "comments": [], "shared_with": [],
        }
    all_leads = [l for l in LEADS if l != owner_lead]

    if request.method == "POST" and is_owner:
        action = request.form.get("action")
        if action == "save_review":
            for cat in RATING_CATEGORIES:
                try:
                    review_data[cat] = int(request.form.get(cat, 0))
                except ValueError:
                    review_data[cat] = 0
            review_data["notes"] = request.form.get("notes", "")
            review_data["status"] = request.form.get("status", "Pending")
            save_review(owner_lead, emp_name, review_data)
            flash("Review saved.")
            return redirect(url_for("review", emp_name=emp_name))
        elif action == "share":
            review_data["shared_with"] = request.form.getlist("share_with")
            save_review(owner_lead, emp_name, review_data)
            flash("Sharing updated.")
            return redirect(url_for("review", emp_name=emp_name))

    notes_text = review_data.get("notes", "") or ""

    return render_template(
        "review.html",
        lead=lead, emp_name=emp_name, emp_info=info,
        review=review_data, is_owner=is_owner, owner_lead=owner_lead,
        categories=RATING_CATEGORIES,
        notes_text=notes_text,
        all_leads=all_leads,
    )


# ── JSON APIs ─────────────────────────────────────────────────────────────────

@app.route("/api/save_review", methods=["POST"])
def api_save_review():
    if "lead" not in session:
        return jsonify({"ok": False}), 401
    if session.get("role") == "director":
        return jsonify({"ok": False, "error": "View-only"}), 403
    lead = session["lead"]
    try:
        data = request.get_json()
        emp_name = data.get("emp_name")
        emp_info, _ = load_org()
        owner_lead = emp_info.get(emp_name, {}).get("lead")
        if lead != owner_lead:
            return jsonify({"ok": False, "error": "Not authorized"}), 403
        review_data = load_review(lead, emp_name) or {}
        for cat in RATING_CATEGORIES:
            if cat in data:
                review_data[cat] = data[cat]
        if "notes" in data:
            review_data["notes"] = _clean_notes(data["notes"])
        if "lead_comments" in data:
            review_data["lead_comments"] = data["lead_comments"]
        if "status" in data:
            review_data["status"] = data["status"]
        review_data.setdefault("employee", emp_name)
        save_review(lead, emp_name, review_data)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/add_comment", methods=["POST"])
def api_add_comment():
    if "lead" not in session:
        return jsonify({"ok": False}), 401
    if session.get("role") == "director":
        return jsonify({"ok": False, "error": "View-only"}), 403
    lead = session["lead"]
    try:
        data = request.get_json()
        emp_name = data.get("emp_name")
        comment_text = data.get("comment", "").strip()
        emp_info, _ = load_org()
        owner_lead = emp_info.get(emp_name, {}).get("lead")
        is_owner = (lead == owner_lead)
        if not is_owner:
            rev_check = load_review(owner_lead, emp_name)
            shared_with_lc = [s.lower() for s in (rev_check or {}).get("shared_with", [])]
            if not rev_check or lead not in shared_with_lc:
                return jsonify({"ok": False, "error": "Access denied"}), 403
        review_data = load_review(owner_lead, emp_name) or {}
        comments = review_data.get("comments", [])
        if isinstance(comments, str):
            try:
                comments = json.loads(comments)
            except Exception:
                comments = []
        new_comment = {
            "author": lead,
            "text": comment_text,
            "time": datetime.now().strftime("%b %d, %Y %H:%M"),
        }
        comments.append(new_comment)
        review_data["comments"] = comments
        review_data.setdefault("employee", emp_name)
        save_review(owner_lead, emp_name, review_data)
        return jsonify({"ok": True, "comment": new_comment})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/update_sharing", methods=["POST"])
def api_update_sharing():
    if "lead" not in session:
        return jsonify({"ok": False}), 401
    if session.get("role") == "director":
        return jsonify({"ok": False}), 403
    lead = session["lead"]
    try:
        data = request.get_json()
        emp_name = data.get("emp_name")
        emp_info, _ = load_org()
        owner_lead = emp_info.get(emp_name, {}).get("lead")
        if lead != owner_lead:
            return jsonify({"ok": False, "error": "Not authorized"}), 403
        review_data = load_review(lead, emp_name) or {}
        old_shared = set(s.lower() for s in review_data.get("shared_with", []))
        new_shared  = [s.lower() for s in data.get("shared_with", [])]
        review_data["shared_with"] = new_shared
        review_data.setdefault("employee", emp_name)
        save_review(lead, emp_name, review_data)
        removed_leads = old_shared - set(new_shared)
        for new_lead in (set(new_shared) - old_shared):
            add_notification(new_lead, f"{lead.capitalize()} shared {emp_name}'s review with you")
            bust_cache(f"shared_{new_lead}")
        for gone_lead in removed_leads:
            bust_cache(f"shared_{gone_lead}")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/delete_review", methods=["POST"])
def api_delete_review():
    if "lead" not in session:
        return jsonify({"ok": False}), 401
    if session.get("role") == "director":
        return jsonify({"ok": False, "error": "View-only"}), 403
    lead = session["lead"]
    try:
        data = request.get_json()
        emp_name = data.get("emp_name")
        emp_info, _ = load_org()
        owner_lead = emp_info.get(emp_name, {}).get("lead")
        if lead != owner_lead:
            return jsonify({"ok": False, "error": "Not authorized"}), 403
        sh = get_spreadsheet()
        ws = _get_ws_ci(sh, lead)
        all_data = ws.get_all_values()
        for i, row in enumerate(all_data[1:], start=2):
            if row and row[0] == emp_name:
                ws.delete_rows(i)
                return jsonify({"ok": True})
        return jsonify({"ok": True})  # already gone
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/change_password", methods=["POST"])
def api_change_password():
    if "lead" not in session:
        return jsonify({"ok": False}), 401
    data = request.get_json()
    current_pw = data.get("current_password", "")
    new_pw = data.get("new_password", "")
    if not new_pw or len(new_pw) < 4:
        return jsonify({"ok": False, "error": "New password must be at least 4 characters"}), 400
    lead = session["lead"]
    matched_key = next((k for k in USERS if k.lower() == lead.lower()), None)
    if not matched_key:
        return jsonify({"ok": False, "error": "User not found"}), 404
    if USERS[matched_key]["password"] != current_pw:
        return jsonify({"ok": False, "error": "Current password is incorrect"}), 400
    USERS[matched_key]["password"] = new_pw
    try:
        _update_password_in_file(matched_key, new_pw)
    except Exception:
        pass
    return jsonify({"ok": True})


@app.route("/resources")
def resources():
    if "lead" not in session:
        return redirect(url_for("login"))
    if session.get("role") == "director":
        flash("Directors manage resources via the dashboard.")
        return redirect(url_for("dashboard"))
    lead = session["lead"]
    emp_info, lead_emps = load_org()
    employees = [
        {"name": emp, "role": emp_info.get(emp, {}).get("role", "Employee")}
        for emp in lead_emps.get(lead, [])
    ]
    return render_template("resources.html", lead=lead, employees=employees)


@app.route("/api/add_resource", methods=["POST"])
def api_add_resource():
    if "lead" not in session:
        return jsonify({"ok": False}), 401
    if session.get("role") == "director":
        return jsonify({"ok": False, "error": "Not authorized"}), 403
    lead = session["lead"]
    data = request.get_json()
    emp_name = (data.get("emp_name") or "").strip()
    emp_role = (data.get("emp_role") or "Employee").strip()
    if not emp_name:
        return jsonify({"ok": False, "error": "Name is required"}), 400
    ok, msg = add_employee_to_org(lead, emp_name, emp_role)
    if not ok:
        return jsonify({"ok": False, "error": msg}), 400
    for director in DIRECTORS:
        add_notification(director.lower(), f"{lead.capitalize()} added {emp_name} ({emp_role}) to their team")
    return jsonify({"ok": True})


@app.route("/api/remove_resource", methods=["POST"])
def api_remove_resource():
    if "lead" not in session:
        return jsonify({"ok": False}), 401
    if session.get("role") == "director":
        return jsonify({"ok": False, "error": "Not authorized"}), 403
    lead = session["lead"]
    data = request.get_json()
    emp_name = (data.get("emp_name") or "").strip()
    ok, msg = remove_employee_from_org(lead, emp_name)
    if not ok:
        return jsonify({"ok": False, "error": msg}), 400
    return jsonify({"ok": True})


@app.route("/notifications")
def notifications():
    if "lead" not in session:
        return redirect(url_for("login"))
    lead = session["lead"]
    notifs = load_notifications(lead)
    notifs_sorted = sorted(notifs, key=lambda x: x.get("timestamp", ""), reverse=True)
    for n in notifs_sorted:
        if not n.get("link"):
            n["link"] = _parse_notif_link(n.get("message", ""))
    mark_notifications_read(lead)
    return render_template("notifications.html", notifications=notifs_sorted)


@app.route("/api/notifications")
def api_notifications():
    if "lead" not in session:
        return jsonify({"ok": False}), 401
    lead = session["lead"]
    notifs = load_notifications(lead)
    unread = sum(1 for n in notifs if n.get("is_read", "false") == "false")
    return jsonify({"ok": True, "unread": unread})


@app.route("/healthz")
def healthz():
    errors = []
    info = {}
    if not os.environ.get("GOOGLE_CREDENTIALS") and not os.path.exists("credentials.json"):
        errors.append("GOOGLE_CREDENTIALS env var not set and credentials.json not found")
    if not os.environ.get("SPREADSHEET_ID"):
        errors.append("SPREADSHEET_ID env var not set")
    try:
        get_client()
        info["auth"] = "ok"
    except Exception as e:
        errors.append(f"Google auth failed: {e}")
    try:
        sh = get_spreadsheet()
        tabs = [ws.title for ws in sh.worksheets()]
        info["spreadsheet_tabs"] = tabs
    except Exception as e:
        errors.append(f"Spreadsheet open failed: {e}")
    try:
        emp_info, lead_emps = load_org()
        info["org_employees"] = len(emp_info)
        info["org_leads"] = list(lead_emps.keys())
    except Exception as e:
        errors.append(f"load_org failed: {e}")
    if errors:
        return jsonify({"ok": False, "errors": errors, "info": info}), 500
    return jsonify({"ok": True, "info": info})


if __name__ == "__main__":
    app.run(debug=True, port=5000)
